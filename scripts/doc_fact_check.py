#!/usr/bin/env python3
"""
文档表述准确性核对脚本（v3 — 复合表述分解增强版）。

工作流程：
1. 将reference_dir下的所有docx/doc文件转换为txt（使用pandoc，.doc用textutil回退）
2. 将目标文档docx转换为txt并提取所有定性和定量表述
3. 将复合表述按分隔符拆解为多个子命题，在全部参考文档中逐一检索
4. 增加实体覆盖检查：核实表述中所有专有名词是否在参考文档中出现
5. 标记状态并输出到Excel

v3 新增能力：
- 复合表述分解：将"A入选X，B获批Y，C实现Z"拆分为独立子命题分别验证
- 实体覆盖检查：提取句内专有名词（引号、书名号内文本），检查与参考文档的一致性
- 子命题独立性评分：任一子命题无出处则整体降级
- .doc 文件自动回退到 textutil 转换

用法:
  python3 doc_fact_check.py <目标文档.docx> <参考文档目录> [输出Excel路径]
  python3 doc_fact_check.py --regenerate <json路径> <输出Excel路径>

输出：
- txt_output/checklist_result.json         第一轮初步结果
- txt_output/checklist_reverse_check.json  反向验证辅助报告
- 核对清单.xlsx                            完整Excel核对清单（5个工作表）
"""

import sys
import os
import json
import glob
import re
import subprocess
from pathlib import Path

try:
    import yaml
except ImportError:
    yaml = None  # 未安装 PyYAML 时回退到内置默认类别词

# ①: 文档内一致性检查默认关闭（启发式误报率过高，详见 Step 3.5 注释）
ENABLE_INTRA_DOC_CHECK = False


# ── 工具函数 ──────────────────────────────────────────────

def convert_docx_to_txt(docx_path, txt_dir):
    """使用pandoc将docx/doc文件转换为txt，.doc文件自动用textutil回退"""
    os.makedirs(txt_dir, exist_ok=True)
    docx_path = os.path.abspath(docx_path)
    basename = os.path.splitext(os.path.basename(docx_path))[0]
    txt_path = os.path.join(txt_dir, basename + ".txt")

    if os.path.exists(txt_path):
        print(f"  [跳过] {basename}.txt 已存在")
        return txt_path

    ext = os.path.splitext(docx_path)[1].lower()
    if ext == '.doc':
        # .doc 文件用 textutil 先转 .docx 再转 txt
        import tempfile
        tmp_docx = os.path.join(tempfile.gettempdir(), basename + "_temp.docx")
        try:
            subprocess.run(['textutil', '-convert', 'docx', docx_path, '-output', tmp_docx],
                           capture_output=True, text=True, timeout=60, check=True)
            docx_path = tmp_docx
        except Exception as e:
            print(f"  [失败] {basename}: .doc转换失败 - {e}")
            return txt_path

    try:
        result = subprocess.run(
            ["pandoc", docx_path, "-t", "plain", "--wrap=none", "-o", txt_path],
            capture_output=True, text=True, timeout=120
        )
        if result.returncode == 0:
            print(f"  [完成] {basename}.txt")
        else:
            print(f"  [失败] {basename}: {result.stderr[:200]}")
    except Exception as e:
        print(f"  [错误] {basename}: {e}")
    return txt_path


def find_doc_files(ref_dir):
    """在目录中查找所有 .docx 和 .doc 文件（递归）"""
    patterns = ["*.docx", "*.doc"]
    files = []
    for pat in patterns:
        found = glob.glob(os.path.join(ref_dir, pat))
        if not found:
            found = glob.glob(os.path.join(ref_dir, "**", pat), recursive=True)
        files.extend(found)
    return sorted(set(files))


def find_txt_refs(ref_dir):
    """在目录中递归查找所有 .txt 参考文件（已转换好的，无需再经 pandoc）"""
    found = glob.glob(os.path.join(ref_dir, "*.txt"))
    found += glob.glob(os.path.join(ref_dir, "**", "*.txt"), recursive=True)
    return sorted(set(found))


# ── 表述提取 ──────────────────────────────────────────────

_NUMERIC_SUFFIXES = (
    r'[万亿]?[个项篇章人次所支部门家国种届轮次场类]'
    r'|[万亿]?%'
    r'|[万亿]?元'
    r'|[万亿]?倍'
    r'|[万亿]?(?:平方米|平米|m\u00b2)'
    r'|[年月日]'
)


def extract_all_numbers(text):
    """从文本中提取所有「数字+单位」组合"""
    pat = re.compile(r'([\d.]+)\s*(' + _NUMERIC_SUFFIXES + r')')
    results = []
    for m in pat.finditer(text):
        results.append(m.group(0).strip())
    return sorted(set(results), key=lambda x: len(x), reverse=True)



def _get_category_patterns(config_path=None):
    """
    加载通用类别词配置，生成正则模式列表。

    配置加载优先级：
    1. 显式指定的 config_path
    2. 脚本同目录下的 entity_config.yaml
    3. 内置默认列表（跨领域通用）

    返回：list of compiled regex patterns
    """
    # 内置默认类别词（跨领域通用，不依赖任何特定领域）
    default_categories = [
        # 任务/任务类型
        '任务', '试点', '改革', '项目', '工程',
        '计划', '行动', '方案', '专项', '示范',
        '建设', '试点单位',
        # 机构/组织
        '机构', '中心', '平台', '基地', '联盟',
        '委员会', '办公室', '部门',
        '学院', '医院', '学校', '大学',
        '研究院', '研究所', '实验室',
        # 体系/模式
        '体系', '模式', '机制', '制度',
        '框架', '范式', '格局', '生态',
        # 文档/标准
        '规划', '报告', '标准', '规范',
        '文件', '办法', '规定',
        # 人员/团队
        '团队', '小组', '队伍',
        # 基础设施/技术
        '系统', '网络', '设施', '设备',
        '模型', '工具',
        # 活动/事件
        '论坛', '会议', '大赛', '活动',
        '峰会', '展览', '培训',
    ]

    # 尝试加载外部配置
    loaded_categories = None
    if config_path and os.path.exists(config_path):
        try:
            if yaml:
                with open(config_path, 'r', encoding='utf-8') as f:
                    loaded_categories = yaml.safe_load(f)
        except Exception:
            pass

    # 如果没指定 config_path，尝试脚本同目录下的默认配置
    if loaded_categories is None:
        default_config_path = os.path.join(os.path.dirname(__file__), 'entity_config.yaml')
        if os.path.exists(default_config_path):
            try:
                if yaml:
                    with open(default_config_path, 'r', encoding='utf-8') as f:
                        loaded_categories = yaml.safe_load(f)
            except Exception:
                pass

    # 从配置中提取所有类别词
    all_suffixes = set(default_categories)
    if loaded_categories:
        for section_name, section_content in loaded_categories.items():
            if isinstance(section_content, dict):
                for sub_name, words in section_content.items():
                    if isinstance(words, list):
                        for w in words:
                            w = w.strip()
                            if len(w) >= 2:
                                all_suffixes.add(w)
            elif isinstance(section_content, list):
                for w in section_content:
                    w = w.strip()
                    if len(w) >= 2:
                        all_suffixes.add(w)

    # 按长度降序排列（长词优先匹配）
    sorted_suffixes = sorted(all_suffixes, key=len, reverse=True)

    # 生成正则模式
    suffix_pattern = r'[\u4e00-\u9fff]{2,30}(?:' + '|'.join(re.escape(s) for s in sorted_suffixes) + r')'
    return [re.compile(suffix_pattern)]

def extract_named_entities(text, category_patterns=None):
    """
    从文本中提取专有名词实体，用于实体覆盖检查。
    返回实体列表，按长度降序。

    提取来源：
    - 中文双引号 \u201c\u201d 内文本
    - 书名号《》内文本
    - 英文引号 "..." 内文本
    - 中文单引号 '' 内文本
    - 特定模式：XX实验室、XX中心、XX项目、XX平台、XX团队、XX学院
    """
    entities = []

    # 中文双引号
    for q in re.findall(r'\u201c([^\u201c\u201d]{2,50})\u201d', text):
        entities.append(q.strip())

    # 英文引号
    for q in re.findall(r'"([^"]{3,50})"', text):
        entities.append(q.strip())

    # 书名号
    for q in re.findall(r'\u300a([^\u300a\u300b]{2,50})\u300b', text):
        entities.append(q.strip())

    # 通用类别词模式（跨领域）
    if category_patterns is None:
        category_patterns = _get_category_patterns()
    for pat in category_patterns:
        for m in pat.findall(text):
            entities.append(m.strip())

    # 过滤假实体：含句内标点的是整句话而非专名
    PUNCT_FILTER = re.compile(r'[，、；。！？,;]')
    entities = [e for e in entities if not PUNCT_FILTER.search(e)]

    # 去重，按长度降序
    seen = set()
    result = []
    for e in entities:
        if e not in seen:
            seen.add(e)
            result.append(e)
    result.sort(key=len, reverse=True)
    return result

def check_intra_document_consistency(items):
    """
    文档内一致性检查。

    当同一目标文档中多个条目提到同一关键实体，
    但该实体的关联属性不一致时，标记为潜在不一致。

    检查逻辑：
    1. 收集所有条目中提取到的实体
    2. 找到在多个条目中出现的"共享实体"
    3. 对每个共享实体，比较其关联的"属性词"
       （属性词 = 该条目中紧邻该实体的动词/名词）
    4. 如果不同条目的属性词差异较大，发出警告

    返回：list of (shared_entity, item_a_seq, attr_a, item_b_seq, attr_b)
    """
    if not items:
        return []

    # 构建：实体 → [(序号, 表述内容, 属性词), ...]
    entity_map = {}
    for idx, item in enumerate(items):
        entities = item.get("命中实体", [])
        content = item.get("表述内容", "")
        for ent in entities:
            if ent not in entity_map:
                entity_map[ent] = []
            # 提取该条目中与实体共现的属性词
            # 属性词 = 位于实体前后20字范围内的实词
            attr_words = _extract_attr_words(content, ent)
            entity_map[ent].append((idx + 1, content[:80], attr_words))

    # 筛选：出现在2个及以上条目中的实体
    conflicts = []
    for ent, occurrences in entity_map.items():
        if len(occurrences) < 2:
            continue
        # 比较各条目中实体的属性词集合
        attr_sets = [set(o[2]) for o in occurrences]
        # 如果属性词集合差距大，可能是不一致
        for i in range(len(attr_sets)):
            for j in range(i + 1, len(attr_sets)):
                common = attr_sets[i] & attr_sets[j]
                union = attr_sets[i] | attr_sets[j]
                if union and len(common) / len(union) < 0.3:
                    # 属性词交集不足30%，可能有矛盾
                    conflicts.append((
                        ent,
                        occurrences[i][0], list(attr_sets[i])[:5],
                        occurrences[j][0], list(attr_sets[j])[:5]
                    ))

    return conflicts


def _extract_attr_words(text, entity, window=20):
    """
    从文本中提取与实体相关的属性词。
    属性词 = 实体前后window字范围内的其他实体/名词。
    """
    idx = text.find(entity)
    if idx == -1:
        return []
    start = max(0, idx - window)
    end = min(len(text), idx + len(entity) + window)
    context = text[start:end]
    # 提取上下文中的其他名词性成分（长度3-15的连续汉字，排除实体本身）
    words = re.findall(r'[\u4e00-\u9fff]{3,15}', context)
    words = [w for w in words if w != entity and w not in text[idx:idx+len(entity)]]
    return words



def decompose_statement(text):
    """
    将复合表述按中文标点拆分为子命题。

    分隔符：分号、逗号（但保留顿号分隔的并列项不拆分）
    策略：
    1. 先按 ； 拆分（强分隔）
    2. 再按 ， 拆分（弱分隔，但保留明显列举的短项）

    返回：[(子命题文本, 是否独立验证), ...]
    独立验证 = True 表示这个子命题包含独立可验证的信息
    """
    # 先按分号拆分
    parts = re.split(r'[；]', text)
    sub_claims = []

    for part in parts:
        part = part.strip()
        if len(part) < 8:
            continue
        # 对逗号分隔的长子命题进一步拆分
        comma_parts = re.split(r'[，]', part)
        for cp in comma_parts:
            cp = cp.strip()
            if len(cp) >= 8:
                # 判断是否含可验证信息
                has_numbers = bool(extract_all_numbers(cp))
                has_entities = bool(extract_named_entities(cp))
                has_action = bool(re.search(
                    r'(?:入选|获批|荣获|获得|新增|组建|成立|牵头|实施|推进|完成|突破|增长|提升|达到|建成|上线|发布|通过|授牌|签约|设立|启用|开创|实现)',
                    cp))
                sub_claims.append({
                    "text": cp,
                    "independent": has_numbers or has_entities or has_action,
                    "numbers": extract_all_numbers(cp),
                    "entities": extract_named_entities(cp),
                })

    return sub_claims


def extract_statements(main_txt_path):
    """从目标文档txt中提取所有定性和定量表述"""
    with open(main_txt_path, 'r', encoding='utf-8') as f:
        content = f.read()

    statements = []
    raw_parts = re.split(r'[。；\n]+', content)
    raw_parts = [p.strip() for p in raw_parts if len(p.strip()) > 10]

    # 预加载类别词模式（只需加载一次）
    category_patterns = _get_category_patterns()

    for part in raw_parts:
        if re.match(r'^[\s\d\.\u3001\uff09\)]+$', part):
            continue
        if len(part) < 10:
            continue

        numbers = extract_all_numbers(part)
        entities = extract_named_entities(part, category_patterns)
        sub_claims = decompose_statement(part)
        stype = "定量" if numbers else "定性"

        statements.append({
            "类型": stype,
            "表述内容": part,
            "状态": "\u2717 \u672a\u627e\u5230",
            "出处": "\u2014",
            "原文片段": "所有参照文档均未出现此表述/数据",
            "匹配关键词": "",
            "命中数字": numbers,
            "命中实体": entities,
            "子命题": sub_claims,
            "匹配上下文简述": "",
            "反向验证警告": "",
        })

    # 去重（④: 用完整表述做键，避免开头30字相同的不同句被误删而漏检）
    seen = set()
    unique = []
    for s in statements:
        key = s["表述内容"]
        if key not in seen:
            seen.add(key)
            unique.append(s)
    return unique


# ── 关键词提取与匹配 ──────────────────────────────────────

def extract_keywords(text):
    """从表述文本中提取用于搜索的关键词，按优先级返回"""
    keywords = []

    # 第一优先：引号内的专有名词（逐对匹配，避免跨类型误匹配）
    quote_pairs = [
        (r'\u201c', r'\u201d'),   # 中文双引号 ""
        (r'\u2018', r'\u2019'),   # 中文单引号 ''
        (r'\u300c', r'\u300d'),   # 直角引号「」
        (r'\u300e', r'\u300f'),   # 直角双引号『』
        (r'\u300a', r'\u300b'),   # 书名号《》
        (r'"',      r'"'),        # ASCII 双引号
    ]
    for open_q, close_q in quote_pairs:
        for m in re.findall(open_q + r'([^' + open_q + close_q + r']{2,})' + close_q, text):
            if 3 <= len(m.strip()) <= 40:
                keywords.append(m.strip())

    # 第二优先：数字+单位组合
    num_matches = re.findall(r'[\d.]+' + _NUMERIC_SUFFIXES, text)
    keywords.extend(n.strip() for n in num_matches if 3 <= len(n.strip()) <= 40)

    # 第三优先：特定动作+名词短语
    action_verbs = r'(?:入选|获批|荣获|获得|新增|组建|成立|牵头|实施|推进|完成|突破|增长|提升|达到|建成|上线|发布|通过|授牌|签约|设立|启用|开创)'
    action_phrases = re.findall(action_verbs + r'[\u4e00-\u9fff]{2,20}', text)
    keywords.extend(a.strip() for a in action_phrases if 3 <= len(a.strip()) <= 40)

    # 第四优先：4~40 字连续中文
    long_chinese = re.findall(r'[\u4e00-\u9fff]{4,40}', text)
    long_chinese.sort(key=len, reverse=True)
    keywords.extend(long_chinese[:5])

    # 去重，按长度降序
    seen = set()
    result = []
    for kw in keywords:
        if kw not in seen:
            seen.add(kw)
            result.append(kw)
    result.sort(key=len, reverse=True)
    fallback = [text[:20]] if text.strip() else []
    return result[:8] if result else fallback


def get_context_snippet(content, keyword, context_width=80):
    """获取关键词在内容中的上下文片段"""
    idx = content.find(keyword)
    if idx == -1:
        return ""
    start = max(0, idx - context_width)
    end = min(len(content), idx + len(keyword) + context_width)
    return content[start:end].replace('\n', ' ')


def build_ref_index(references):
    """构建参考文档索引"""
    index = {}
    for txt_path, txt_content in references:
        short_name = os.path.splitext(os.path.basename(txt_path))[0][:40]
        index[short_name] = (txt_path, txt_content)
    return index


def verify_sub_claims(sub_claims, references):
    """
    独立验证每个子命题在参考文档中的匹配情况。

    返回：
    - match_flags: [bool, ...] 每个子命题是否找到匹配
    - total: 子命题总数
    - matched: 匹配到的子命题数
    - details: list of str 详细说明
    """
    independent_claims = [sc for sc in sub_claims if sc.get("independent")]
    if not independent_claims:
        return [], 0, 0, []

    match_flags = []
    details = []

    # 有锚点的子命题才做独立出处检索：含数字、引号/书名号专名、或成就动词
    _ACHIEVEMENT_VERBS = re.compile(
        r'(获批|入选|获评|承担|建立|成立|设立|推出|完成|实现|'
        r'签署|出版|发布|获得|荣获|入围|评为|被授予|被遴选|牵头)'
    )
    _HAS_ANCHOR = re.compile(
        r'[\d０-９]|[《""“”]|'
        r'(获批|入选|获评|承担|建立|成立|设立|推出|完成|实现|'
        r'签署|出版|发布|获得|荣获|入围|评为|被授予|被遴选|牵头)'
    )

    for sc in independent_claims:
        # 无锚点的论述性子命题跳过，不产生"无独立出处"警告
        if not _HAS_ANCHOR.search(sc["text"]):
            match_flags.append(True)   # 视为通过，不拉低覆盖率
            continue

        sub_kws = extract_keywords(sc["text"])
        sub_found = False

        for kw in sub_kws:
            for _, ref_content in references:
                if kw in ref_content:
                    sub_found = True
                    break
            if sub_found:
                break

        match_flags.append(sub_found)
        if not sub_found:
            details.append(f"子命题「{sc['text'][:50]}」无独立出处")

    total = len(independent_claims)
    matched = sum(match_flags)
    return match_flags, total, matched, details


_ORDINAL_PREFIX = re.compile(r'^[一二三四五六七八九十][是、]')
_DIYIJIEDUAN_PREFIX = re.compile(r'^第[一二三四五六七八九十\d]阶')  # 第X阶段（不过滤第X批）
_CONJUNCTION_PREFIX = re.compile(
    r'^(?:因此|所以|而且|并且|但是|然而|同时|此外|另外|不仅|其次|回顾|纵观|综上)')
# 动词/介词/副词开头的短语 → 大概率是句子片段，不是具名实体
_VERB_PREFIX = re.compile(
    r'^(?:构建|探索|推动|推进|实施|建设|完善|深化|加强|建立|持续|全面|系统'
    r'|完成|开创|优化|整合|强化|提升|发展|创新|落实|推行|健全|开展'
    r'|形成|确保|聚焦|坚持|促进|统筹|围绕|围绕|为了|为各|为更)')
_PRONOUN_PATTERNS = ('我们', '他们', '你们', '我校', '本校')
_TITLE_CHARS = re.compile(r'[《》〈〉「」【】《-』]')
# 语法虚词/助词开头 → 句子尾缀或介词短语，不是具名实体
_PARTICLE_PREFIX = re.compile(r'^[等与个的既]')


def _is_valid_entity(entity: str) -> bool:
    """
    判断 entity 是否是有意义的具名实体（可用于覆盖率和共现验证）。
    过滤掉 LLM 误提取的整句片段：
    - 含人称代词
    - 以序号/阶段描述起头
    - 以连接词/动词/介词起头
    - 以语法虚词（等/与/个/的/既）起头
    - 含「不是」的否定谓语结构（句子片段）
    - 超长（≥20字）且不含书名号/标题符号
    """
    if not entity or not entity.strip():
        return False
    if any(p in entity for p in _PRONOUN_PATTERNS):
        return False
    if _ORDINAL_PREFIX.match(entity):
        return False
    if _DIYIJIEDUAN_PREFIX.match(entity):
        return False
    if _CONJUNCTION_PREFIX.match(entity):
        return False
    if _VERB_PREFIX.match(entity):
        return False
    if _PARTICLE_PREFIX.match(entity):
        return False
    if '不是' in entity:
        return False
    # 超长且无标题符号 → 大概率是句子
    if len(entity) >= 20 and not _TITLE_CHARS.search(entity):
        return False
    return True


def verify_entity_coverage(entities, all_ref_contents):
    """
    实体覆盖检查：目标表述中的专有名词，在所有参考文档中出现了多少。

    all_ref_contents: list of (ref_name, ref_content) tuples
    返回：(覆盖率 %, 未覆盖实体列表)
    """
    if not entities:
        return 100.0, []

    uncovered = []
    for entity in entities:
        found = any(entity in content for _, content in all_ref_contents)
        if not found:
            uncovered.append(entity)

    coverage = (len(entities) - len(uncovered)) / len(entities) * 100
    return coverage, uncovered


def verify_entity_context_cooccurrence(entities, matched_kw, references, context_radius=300):
    """
    v3.1 新增：实体-语境共现验证。

    仅检查实体字符串是否存在是不够的——实体必须与匹配关键词出现在同一语境中。
    例如：目标文档说"金融大模型实验室入选两重项目"，
    关键词"两重"在参考文档中匹配了，实体"金融大模型实验室"也存在于参考文档中，
    但它们在参考文档中从未在同一段落出现——这是张冠李戴。

    参数：
    - entities: 目标表述中的专有名词列表
    - matched_kw: 触发匹配的关键词
    - references: list of (ref_name, ref_content)
    - context_radius: 共现窗口大小（字符数）

    返回：
    - misattributed: list of (entity, ref_name_where_found_but_not_near_kw)
      实体在参考文档中存在但不在关键词附近的列表
    - cooccurring: list of entity 与关键词共现的实体
    """
    if not entities or not matched_kw:
        return [], list(entities)

    misattributed = []
    cooccurring = []

    for entity in entities:
        if entity == matched_kw:
            cooccurring.append(entity)
            continue

        entity_found_anywhere = False
        entity_near_kw = False
        found_in_ref = ""

        for ref_name, ref_content in references:
            if entity in ref_content:
                entity_found_anywhere = True
                if not found_in_ref:
                    # 用文件名（去路径和扩展名）作为标识
                    found_in_ref = os.path.splitext(os.path.basename(ref_name))[0]

            # 检查实体是否与关键词在同一语境中共现
            if entity in ref_content and matched_kw in ref_content:
                # 找到实体和关键词在参考文档中最近的位置
                kw_positions = [m.start() for m in re.finditer(re.escape(matched_kw), ref_content)]
                entity_positions = [m.start() for m in re.finditer(re.escape(entity), ref_content)]
                for ep in entity_positions:
                    for kp in kw_positions:
                        if abs(ep - kp) <= context_radius:
                            entity_near_kw = True
                            break
                    if entity_near_kw:
                        break

        if entity_near_kw:
            cooccurring.append(entity)
        elif entity_found_anywhere:
            # 实体存在但不与关键词共现——张冠李戴风险
            misattributed.append((entity, found_in_ref))
        else:
            # 实体完全不存在——已有 verify_entity_coverage 处理
            cooccurring.append(entity)  # 不在 misattributed 里，由 coverage 负责

    return misattributed, cooccurring


def search_in_reference(texts, references):
    """
    增强版全文检索：
    - Stage 2A: 优先用 retrieval.py（trigram/BM25）检索排序候选，回退到子串匹配
    - 找到首个关键词匹配后继续搜全部参考文档做实体覆盖 + 数字验证
    - 实体覆盖检查跨全部参考文档
    - 子命题独立验证
    - 反向验证标记
    """
    index = build_ref_index(references)
    total = len(texts)
    last_pct = 0

    # Stage 2A: 构建 retrieval 索引（一次性，复用全部 statements）
    _ret_index = None
    _ret_mod = None
    try:
        import retrieval as _ret_mod
        ref_contents = {name: content for name, (_, content) in index.items()}
        _ret_index = _ret_mod.build_index(ref_contents)
    except Exception:
        pass  # 回退到子串匹配

    for idx, item in enumerate(texts):
        # 进度指示
        pct = (idx + 1) * 100 // total
        if pct >= last_pct + 10:
            print(f"  检索进度: {idx+1}/{total} ({pct}%)")
            last_pct = pct

        query = item["表述内容"]
        keywords = extract_keywords(query)
        all_numbers = item.get("命中数字", [])
        # 过滤 LLM 误提取的噪声实体（人称代词/序号/连接词开头的句子片段）
        all_entities = [e for e in item.get("命中实体", []) if _is_valid_entity(e)]
        sub_claims = item.get("子命题", [])

        found_any = False
        matched_kw = ""
        matched_ref = ""
        matched_content = ""

        # 第一遍：找到首个匹配（确定出处和原文片段）
        # Stage 2A: 先用 retrieval 排序候选，再从候选中挑第一个含关键词的
        if _ret_index is not None and _ret_mod is not None:
            try:
                candidates = _ret_mod.retrieve(
                    query, all_numbers, all_entities, _ret_index, top_k=5)
                # 始终保存候选证据：即使关键词未命中，LLM 判定层也能用
                if candidates:
                    item["候选证据"] = candidates[:5]
                for cand in candidates:
                    ref_name = cand["sourcePath"]
                    if ref_name not in index:
                        continue
                    _, ref_content = index[ref_name]
                    # 只接受关键词在 chunk 文本中实际出现的候选，防止弱证据假正例
                    kw_hit = next(
                        (kw for kw in keywords if kw in cand["text"]),
                        None)
                    if kw_hit is None:
                        continue  # 此候选无关键词命中，跳过
                    snippet = get_context_snippet(ref_content, kw_hit, 80)
                    item["状态"] = "\u2713 \u5df2\u786e\u8ba4"
                    item["出处"] = ref_name
                    item["原文片段"] = f"{ref_name}: ...{snippet}..."
                    item["匹配关键词"] = kw_hit
                    item["匹配上下文简述"] = snippet[:120]
                    matched_kw = kw_hit
                    matched_ref = ref_name
                    matched_content = ref_content
                    found_any = True
                    break
            except Exception:
                pass  # 降级到子串

        # 回退：原子串匹配
        if not found_any:
            for kw in keywords:
                for ref_name, (ref_path, ref_content) in index.items():
                    if kw in ref_content:
                        snippet = get_context_snippet(ref_content, kw, 80)
                        item["状态"] = "\u2713 \u5df2\u786e\u8ba4"
                        item["出处"] = ref_name
                        item["原文片段"] = f"{ref_name}: ...{snippet}..."
                        item["匹配关键词"] = kw
                        item["匹配上下文简述"] = snippet[:120]
                        matched_kw = kw
                        matched_ref = ref_name
                        matched_content = ref_content
                        found_any = True
                        break
                if found_any:
                    break

        if not found_any:
            continue

        # ── 反向验证标记 ──
        warnings = []

        # === v3: 子命题独立性验证 ===
        mf, total_sp, matched_sp, sub_details = verify_sub_claims(sub_claims, references)
        if total_sp > 1:
            if matched_sp < total_sp:
                if "\u25b3" not in item["状态"]:
                    item["状态"] = "\u25b3 \u90e8\u5206\u5339\u914d"
                sub_warning = f"复合表述含{total_sp}个可独立验证子命题，{matched_sp}个匹配({matched_sp/total_sp*100:.0f}%)，" + "; ".join(sub_details)
                warnings.append(sub_warning)

        # === v3: 实体覆盖检查（跨全部参考文档） ===
        if all_entities:
            coverage, uncovered = verify_entity_coverage(all_entities, references)
            if coverage < 100:
                entity_warning = (
                    f"实体覆盖率{coverage:.0f}%，未在任何参考文档中找到的实体: {uncovered}。"
                    f"可能将不同实体张冠李戴"
                )
                warnings.append(entity_warning)
                if coverage < 50 and len(all_entities) >= 2:
                    if "\u2713" in item["状态"]:
                        item["状态"] = "\u25b3 \u6570\u636e\u4e0d\u4e00\u81f4"

            # === v3.1: 实体-语境共现验证 ===
            # 「政策依据：」行是多文件并列引用，跨文件引用合法，跳过共现检查
            is_policy_ref = bool(re.match(r'^[-\s]*政策依据[：:]', query))
            if matched_kw and not is_policy_ref:
                misattributed, cooccurring = verify_entity_context_cooccurrence(
                    all_entities, matched_kw, references
                )
                if misattributed:
                    # 有实体存在于参考文档中，但不与匹配关键词共现
                    mis_names = [f"{e}(仅见于{ref[:30]}，非'{matched_kw}'语境)"
                                 for e, ref in misattributed]
                    entity_ctx_warning = (
                        f"⚠ 实体-语境不匹配: {', '.join(mis_names)}。"
                        f"关键词'{matched_kw}'匹配成功，但这些实体在参考文档中"
                        f"从不与'{matched_kw}'在同一语境中出现，可能张冠李戴"
                    )
                    warnings.append(entity_ctx_warning)
                    # 关键实体不共现 → 直接降级为数据不一致
                    # 只要有任何实体被张冠李戴（不少于1个且不是仅关键词本身），就降级
                    non_kw_mis = [(e, r) for e, r in misattributed if e != matched_kw]
                    if "\u2713" in item.get("状态", "") and len(non_kw_mis) >= 1:
                        item["状态"] = "\u25b3 \u6570\u636e\u4e0d\u4e00\u81f4"

        # === 原有检查 ===

        # 数字缺失检查（跨全部参考文档）
        if all_numbers:
            missing = [n for n in all_numbers
                       if not any(n in ref_content for _, ref_content in references)]
            if missing:
                warnings.append(f"部分数字未在参考文档中找到: {missing}")

        # 增长率语境检查
        growth_keywords = re.findall(r'增长|增幅|提高|提升|同比|突破|跃升|攀升|翻番', query)
        if growth_keywords:
            pct_nums = [n for n in all_numbers if n.endswith('%')]
            for pn in pct_nums:
                in_growth_ctx = False
                for _, ref_content in references:
                    if pn in ref_content:
                        idx2 = ref_content.find(pn)
                        ctx = ref_content[max(0, idx2-25):idx2+25]
                        if any(gk in ctx for gk in growth_keywords):
                            in_growth_ctx = True
                            break
                if not in_growth_ctx:
                    warnings.append(f"增长率 {pn} 未在增长语境中出现，需人工核实")

        # 关键词过短检查
        if len(matched_kw) <= 3:
            warnings.append(f"匹配关键词过短({len(matched_kw)}字)，可能为误匹配，需人工确认上下文")

        # 关键词过于通用检查（文档频率替代绝对次数，跨任务自适应）
        total_occurrences = sum(ref_content.count(matched_kw) for _, ref_content in references)
        doc_freq = sum(1 for _, ref_content in references if matched_kw in ref_content)
        total_docs = len(references)
        # 出现在≥60%参考文档中 = 领域核心词，不报通用警告
        is_domain_term = total_docs > 0 and (doc_freq / total_docs) >= 0.6
        if total_occurrences > 20 and not is_domain_term:
            warnings.append(
                f"匹配关键词\u300c{matched_kw}\u300d在参考文档中出现{total_occurrences}次"
                f"（覆盖{doc_freq}/{total_docs}份文档），过于通用，"
                f"请确认匹配到的上下文是否与目标表述一致"
            )

        # === ④: 仅靠过短/高频通用词命中的「假✓」降级为「需核实」===
        # 关键词命中即判✓是主要漏报来源：通用词碰巧出现就误判已确认。
        # 若命中仅依赖短词或高频通用词，且无「数字命中」或「引号专名被覆盖」等强证据佐证，
        # 则降级为需人工核实，避免静默假✓。
        if "✓" in item.get("状态", ""):
            kw_short = len(matched_kw) <= 3
            # 领域核心词不触发降级，只有真正低区分度的通用词才降级
            kw_generic = total_occurrences > 20 and not is_domain_term
            strong_ok = False
            if all_numbers and any(
                    any(n in rc for _, rc in references) for n in all_numbers):
                strong_ok = True
            else:
                quoted = re.findall(r'[“《"]([^”》"]{2,40})[”》"]', query)
                if quoted and any(any(q in rc for _, rc in references) for q in quoted):
                    strong_ok = True
            if (kw_short or kw_generic) and not strong_ok:
                item["状态"] = "△ 需核实"
                reason = "/".join(filter(None, [
                    "过短" if kw_short else "", "高频通用" if kw_generic else ""]))
                warnings.append(
                    f"仅靠{reason}关键词'{matched_kw}'命中、"
                    f"无数字或引号专名佐证，"
                    f"已自动降级为需核实（防止通用词假✓漏报）"
                )

        # === ⑤: 短关键词 + 所有数字均缺失 → "✗ 未找到"（防 △部分匹配 假阳性）===
        # Rule ④ 只覆盖「假✓」，但子命题失败后状态已降为「△ 部分匹配」，
        # 此时若匹配仅靠低区分度短词且声明中的全部数字均不在参考文档，
        # 说明根本无实质证据，应进一步降为「未找到」。
        if "△" in item.get("状态", "") and matched_kw and len(matched_kw) <= 3 and all_numbers:
            all_nums_missing = all(
                not any(n in rc for _, rc in references)
                for n in all_numbers
            )
            if all_nums_missing:
                item["状态"] = "✗ 未找到"
                warnings.append(
                    f"关键词'{matched_kw}'(≤3字)命中但声明中全部数字均不在参考文档，"
                    f"实质无证据，降级为未找到"
                )

        item["反向验证警告"] = "; ".join(warnings) if warnings else ""


# ── Excel 生成 ─────────────────────────────────────────────

def generate_excel(items, output_path):
    """生成5工作表的Excel核对清单"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import PatternFill, Border, Side
    except ImportError:
        print("请先安装 openpyxl: pip install openpyxl")
        return

    wb = Workbook()

    green  = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
    yellow = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    red    = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    orange = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    def _auto_width(ws, widths):
        for col_letter, w in widths.items():
            ws.column_dimensions[col_letter].width = w

    def _color_cell(ws, row, col, status):
        cell = ws.cell(row=row, column=col)
        if "\u2713" in status:
            cell.fill = green
        elif "\u6570\u636e\u4e0d\u4e00\u81f4" in status:
            cell.fill = orange
        elif "\u25b3" in status:
            cell.fill = yellow
        elif "\u672a\u627e\u5230" in status:
            cell.fill = red

    # Sheet 1: 全部核对结果
    ws1 = wb.active
    ws1.title = "全部核对结果"
    headers1 = ["序号", "类型", "表述内容", "核对状态", "出处", "原文片段", "反向验证警告", "命中数字"]
    ws1.append(headers1)
    for i, item in enumerate(items, 1):
        row_data = [
            i, item["类型"], item["表述内容"], item["状态"],
            item["出处"], item["原文片段"],
            item.get("反向验证警告", ""),
            ", ".join(item.get("命中数字", []))
        ]
        ws1.append(row_data)
        _color_cell(ws1, i+1, 4, item["状态"])
        if item.get("反向验证警告", ""):
            ws1.cell(row=i+1, column=7).fill = yellow
        for c in range(1, 9):
            ws1.cell(row=i+1, column=c).border = thin_border
    _auto_width(ws1, {'A': 6, 'B': 8, 'C': 55, 'D': 18, 'E': 40, 'F': 90, 'G': 50, 'H': 30})

    # Sheet 2: 已确认
    ws2 = wb.create_sheet("已确认表述")
    ws2.append(["序号", "类型", "表述内容", "出处", "原文片段", "反向验证警告"])
    for i, item in enumerate((x for x in items if "\u2713" in x.get("状态", "")), 1):
        ws2.append([i, item["类型"], item["表述内容"], item["出处"], item["原文片段"],
                    item.get("反向验证警告", "")])
    _auto_width(ws2, {'A': 6, 'B': 8, 'C': 55, 'D': 40, 'E': 90, 'F': 50})

    # Sheet 3: 待核实
    ws3 = wb.create_sheet("待核实表述")
    ws3.append(["序号", "类型", "表述内容", "核对状态", "备注说明"])
    for i, item in enumerate((x for x in items if "\u672a\u627e\u5230" in x.get("状态", "") or "\u25b3" in x.get("状态", "")), 1):
        ws3.append([i, item["类型"], item["表述内容"], item["状态"], item["原文片段"]])
        _color_cell(ws3, i+1, 4, item["状态"])
    _auto_width(ws3, {'A': 6, 'B': 8, 'C': 55, 'D': 18, 'E': 90})

    # Sheet 4: 反向验证重点关注
    ws4 = wb.create_sheet("反向验证重点关注（第三轮）")
    ws4.append(["序号", "类型", "表述内容", "当前状态", "命中数字", "匹配关键词", "反向验证警告", "人工核查建议"])
    suspicious = [item for item in items if item.get("反向验证警告", "")]
    for i, item in enumerate(suspicious, 1):
        warnings = item["反向验证警告"]
        advice = _build_advice(warnings)
        ws4.append([i, item["类型"], item["表述内容"], item["状态"],
                    ", ".join(item.get("命中数字", [])),
                    item.get("匹配关键词", ""), warnings, advice])
        ws4.cell(row=i+1, column=7).fill = yellow
    _auto_width(ws4, {'A': 6, 'B': 8, 'C': 50, 'D': 16, 'E': 30, 'F': 25, 'G': 55, 'H': 60})

    # Sheet 5: 实体覆盖分析（v3新增）
    ws5 = wb.create_sheet("实体覆盖分析")
    ws5.append(["序号", "表述内容", "提取的实体", "匹配状态", "实体覆盖情况"])
    for i, item in enumerate(items, 1):
        entities = item.get("命中实体", [])
        if entities:
            ws5.append([
                i, item["表述内容"][:80],
                ", ".join(entities),
                item["状态"],
                _build_entity_coverage_note(item.get("反向验证警告", ""))
            ])
            _color_cell(ws5, i+1, 4, item["状态"])
    _auto_width(ws5, {'A': 6, 'B': 55, 'C': 50, 'D': 16, 'E': 50})

    wb.save(output_path)
    print(f"\nExcel已保存: {output_path}")
    print(f"  工作表: {wb.sheetnames}")
    print(f"  其中\u300c反向验证重点关注\u300d工作表包含 {len(suspicious)} 条需第三轮复查的条目")


def _build_advice(warnings):
    """根据警告内容生成人工核查建议"""
    parts = []
    if "过短" in warnings:
        parts.append("\u2460 匹配关键词太短，请读取原文片段确认是否真的对应同一表述")
    if "出现" in warnings and "次" in warnings:
        parts.append("\u2461 关键词过于通用，请检查上下文是否与目标表述一致")
    if "数字未找到" in warnings:
        parts.append("\u2462 部分数字无出处，请核实这些数字的来源")
    if "增长率" in warnings:
        parts.append("\u2463 增长率数字缺少对应语境，请确认增长率是否准确")
    if "子命题" in warnings:
        parts.append("\u2464 复合表述有子命题无出处，请逐个子命题核实")
    if "实体覆盖率" in warnings:
        parts.append("\u2465 实体覆盖率不足，可能存在张冠李戴，请逐名称比对")
    return "; ".join(parts) if parts else "请逐字比对目标表述与原文片段是否一致"


def _build_entity_coverage_note(warnings):
    """从警告中提取实体覆盖相关信息"""
    if not warnings:
        return "\u2713 实体全部匹配"
    if "实体覆盖率" in warnings:
        return warnings
    return "见反向验证警告"


def generate_markdown(items, output_path, doc_name="", ref_count=0):
    """生成 Markdown 核对报告（重要优先：✗ → ⚠△ → ✓）"""
    import datetime

    total   = len(items)
    nf_n    = sum(1 for x in items if "✗" in x.get("状态", ""))
    warn_n  = sum(1 for x in items if x.get("反向验证警告", ""))
    delta_n = sum(1 for x in items if "△" in x.get("状态", ""))
    conf_n  = sum(1 for x in items if "✓" in x.get("状态", ""))

    not_found = [x for x in items if "✗" in x.get("状态", "")]
    flagged   = [x for x in items if x.get("反向验证警告", "") or "△" in x.get("状态", "")]
    confirmed = [x for x in items if "✓" in x.get("状态", "")]

    today = datetime.date.today().strftime("%Y-%m-%d")
    title = doc_name or os.path.splitext(os.path.basename(output_path))[0]

    L = []
    L.append(f"# 核对清单 — {title}")
    L.append(f"生成时间：{today}  参考文档：{ref_count} 份  共 {total} 条")
    L.append("")
    L.append("## 概览")
    L.append(
        f"✗ 未找到：{nf_n}　　"
        f"⚠ 反向验证重点关注：{warn_n}　　"
        f"△ 需关注：{delta_n}　　"
        f"✓ 已确认：{conf_n}"
    )
    L.append("")
    L.append("---")

    # ── Section 1: ✗ 未找到 ──
    L.append("")
    L.append(f"## ✗ 未找到（{len(not_found)} 条）—— 建议人工逐条核查")
    if not_found:
        L.append("")
        L.append("| # | 表述内容 | 命中实体 | 命中数字 |")
        L.append("|---|---------|---------|--------|")
        for i, item in enumerate(not_found, 1):
            text     = item["表述内容"][:80].replace("|", "｜")
            entities = "、".join(item.get("命中实体", [])) or "—"
            numbers  = "、".join(item.get("命中数字", [])) or "—"
            L.append(f"| {i} | {text} | {entities} | {numbers} |")
    else:
        L.append("\n（无）")

    L.append("")
    L.append("---")

    # ── Section 2: ⚠/△ 反向验证重点关注 ──
    L.append("")
    L.append(f"## ⚠ 反向验证重点关注（{len(flagged)} 条）")
    if flagged:
        for i, item in enumerate(flagged, 1):
            status  = item.get("状态", "")
            source  = item.get("出处", "") or "—"
            snippet = item.get("原文片段", "") or ""
            if len(snippet) > 120:
                snippet = snippet[:120] + "…"
            snippet_md = f"`{snippet.replace('`', '')}`" if snippet else "—"
            kw      = item.get("匹配关键词", "") or "—"
            warning = item.get("反向验证警告", "") or "—"
            advice  = _build_advice(warning)
            L.append("")
            L.append(f"### {i} · {status}")
            L.append("")
            L.append("**表述：**  ")
            L.append(item["表述内容"])
            L.append("")
            L.append("| 字段 | 内容 |")
            L.append("|------|------|")
            L.append(f"| 出处 | {source} |")
            L.append(f"| 原文片段 | {snippet_md} |")
            L.append(f"| 匹配关键词 | {kw} |")
            L.append(f"| 警告 | {warning} |")
            L.append(f"| 建议 | {advice} |")
    else:
        L.append("\n（无）")

    L.append("")
    L.append("---")

    # ── Section 3: ✓ 已确认（折叠） ──
    L.append("")
    L.append(f"## ✓ 已确认（{len(confirmed)} 条）")
    if confirmed:
        L.append("")
        L.append("<details><summary>展开查看</summary>")
        L.append("")
        L.append("| # | 表述内容（前 80 字） | 出处 |")
        L.append("|---|------------------|------|")
        for i, item in enumerate(confirmed, 1):
            text      = item["表述内容"][:80].replace("|", "｜")
            source    = item.get("出处", "") or "—"
            warn_flag = " ⚠" if item.get("反向验证警告", "") else ""
            L.append(f"| {i} | {text} | {source}{warn_flag} |")
        L.append("")
        L.append("</details>")
    else:
        L.append("\n（无）")

    content = "\n".join(L) + "\n"
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"\nMarkdown 报告已保存: {output_path}")


# ── 统计 ──────────────────────────────────────────────────

def print_summary(statements, stage="第一轮自动核对"):
    """打印汇总统计"""
    confirmed  = len([s for s in statements if "\u2713" in s["\u72b6\u6001"]])
    not_found  = len([s for s in statements if "\u672a\u627e\u5230" in s["\u72b6\u6001"]])
    partial    = len(statements) - confirmed - not_found
    suspicious = len([s for s in statements if s.get("反向验证警告", "")])

    print(f"\n{'='*60}")
    print(f"{stage} 完成")
    print(f"{'='*60}")
    print(f"总表述数:     {len(statements)}")
    print(f"  \u2713 已确认:   {confirmed}")
    print(f"  \u25b3 需核实:   {partial}")
    print(f"  \u2717 未找到:   {not_found}")
    if suspicious:
        print(f"{'='*60}")
        print(f"\u26a0 反向验证警告: {suspicious} 条表述存在潜在问题，需重点复查")
        print(f"  请在 Markdown 报告\u300c反向验证重点关注\u300d章节逐条复查")
        print(f"{'='*60}")


# ── 入口 ──────────────────────────────────────────────────

def cmd_full_check():
    """运行完整的第一轮自动核对流程"""
    if len(sys.argv) < 3:
        print("用法: python3 doc_fact_check.py <目标文档.docx> <参考文档目录> [输出报告.md] [--excel] [--llm-judge]")
        sys.exit(1)

    args  = [a for a in sys.argv[1:] if not a.startswith("--")]
    flags = [a for a in sys.argv[1:] if a.startswith("--")]
    llm_judge  = "--llm-judge" in flags
    want_excel = "--excel" in flags

    main_docx  = args[0] if len(args) > 0 else ""
    ref_dir    = args[1] if len(args) > 1 else ""
    output_md  = args[2] if len(args) > 2 else ""   # 默认下面计算

    if not main_docx or not ref_dir:
        print("用法: python3 doc_fact_check.py <目标文档.docx> <参考文档目录> [输出报告.md] [--excel] [--llm-judge]")
        sys.exit(1)

    txt_dir = os.path.join(os.path.dirname(main_docx) or ".", "txt_output")
    os.makedirs(txt_dir, exist_ok=True)

    # Step 1: 转换参考文档
    print("=" * 60)
    print("Step 1: 转换参考文档 docx/doc \u2192 txt")
    print("=" * 60)
    ref_files = find_doc_files(ref_dir)
    print(f"找到 {len(ref_files)} 个参考文档")

    ref_texts = {}
    for doc in ref_files:
        txt_path = convert_docx_to_txt(doc, txt_dir)
        with open(txt_path, 'r', encoding='utf-8') as f:
            ref_texts[txt_path] = f.read()
        print(f"  \u2192 {os.path.basename(doc)}")

    # 直接读入已有 .txt 参考文件（无需 pandoc 转换）
    txt_ref_files = find_txt_refs(ref_dir)
    txt_only = [p for p in txt_ref_files if p not in ref_texts]
    if txt_only:
        print(f"  直接读取 {len(txt_only)} 个 .txt 参考文件")
        for txt_path in txt_only:
            try:
                with open(txt_path, 'r', encoding='utf-8') as f:
                    ref_texts[txt_path] = f.read()
            except Exception as e:
                print(f"  [跳过] {os.path.basename(txt_path)}: {e}")

    # Step 2: 提取表述
    print("\n" + "=" * 60)
    print("Step 2: 提取目标文档表述 (v3: 含子命题分解 + 实体提取)")
    print("=" * 60)
    main_txt = convert_docx_to_txt(main_docx, txt_dir)
    statements = extract_statements(main_txt)

    # 统计复合表述
    compound_count = sum(1 for s in statements if len(s.get("子命题", [])) > 1)
    entity_count = sum(1 for s in statements if s.get("命中实体", []))
    print(f"提取到 {len(statements)} 条表述（定性+定量）")
    print(f"  其中复合表述 {compound_count} 条，含专有实体的表述 {entity_count} 条")

    # Step 3: 检索 + 反向验证
    print("\n" + "=" * 60)
    print("Step 3: 全文检索 + 子命题验证 + 实体覆盖检查")
    print("=" * 60)
    ref_list = list(ref_texts.items())
    search_in_reference(statements, ref_list)

    # Step 3.5: 文档内一致性检查（v3.2 新增）
    # ①: 该启发式（±20字窗口内随机汉字串 Jaccard<0.3）与真矛盾基本无关，
    # 在讲话稿等高复现度文本上几乎全是假警告（如"引进来/走出去""一流大学"），
    # 反而淹没真信号、抬高漏看概率，故默认关闭。需要时设 ENABLE_INTRA_DOC_CHECK=True。
    print("\n" + "=" * 60)
    print("Step 3.5: 文档内一致性检查")
    print("=" * 60)
    conflicts = check_intra_document_consistency(statements) if ENABLE_INTRA_DOC_CHECK else []
    if not ENABLE_INTRA_DOC_CHECK:
        print("  已默认禁用（误报率过高）。如需启用：ENABLE_INTRA_DOC_CHECK=True")
    if conflicts:
        print(f"发现 {len(conflicts)} 处潜在不一致：")
        for ent, seq_a, attrs_a, seq_b, attrs_b in conflicts:
            warning = (
                f"文档内不一致：实体「{ent}」在条目{seq_a}中关联"
                f"属性{attrs_a}，在条目{seq_b}中关联"
                f"属性{attrs_b}，可能自相矛盾"
            )
            print(f"  ⚠ {warning}")
            # 将警告写入对应条目
            for s in statements:
                s_idx = statements.index(s)
                if s_idx + 1 == seq_a or s_idx + 1 == seq_b:
                    existing = s.get("反向验证警告", "")
                    if "文档内不一致" not in existing:
                        if existing:
                            s["反向验证警告"] = existing + "; " + warning
                        else:
                            s["反向验证警告"] = warning
    elif ENABLE_INTRA_DOC_CHECK:
        print("  未发现文档内矛盾")
    print()

    # Step 3.8: LLM 判定（--llm-judge 开关）
    if llm_judge:
        print("\n" + "=" * 60)
        print("Step 3.8: LLM 判定层（--llm-judge）")
        print("=" * 60)
        try:
            import adjudicate as _adj
            _adj.adjudicate_batch(statements, verbose=True)
        except ImportError:
            print("  ⚠ adjudicate.py 未找到，跳过 LLM 判定")
    else:
        print("  （LLM 判定已跳过，使用 --llm-judge 开启）")

    # Step 4: 保存中间结果
    print("\n" + "=" * 60)
    print("Step 4: 保存中间结果")
    print("=" * 60)
    json_path = os.path.join(txt_dir, "checklist_result.json")
    with open(json_path, 'w', encoding='utf-8') as f:
        json.dump(statements, f, ensure_ascii=False, indent=2)
    print(f"中间结果已保存: {json_path}")

    reverse_check = [s for s in statements if s.get("反向验证警告", "")]
    rev_path = os.path.join(txt_dir, "checklist_reverse_check.json")
    with open(rev_path, 'w', encoding='utf-8') as f:
        json.dump(reverse_check, f, ensure_ascii=False, indent=2)
    print(f"反向验证报告已保存: {rev_path}")

    # Step 5: 生成 Markdown 报告（默认）；可选 --excel
    print("\n" + "=" * 60)
    print("Step 5: 生成核对报告")
    print("=" * 60)
    doc_basename = os.path.splitext(os.path.basename(main_docx))[0]
    if not output_md:
        output_md = os.path.join(os.path.dirname(main_docx) or ".", f"{doc_basename}_核对清单.md")
    elif not os.path.isabs(output_md):
        output_md = os.path.join(os.path.dirname(main_docx) or ".", output_md)
    generate_markdown(statements, output_md, doc_name=doc_basename, ref_count=len(ref_texts))

    if want_excel:
        output_xlsx = os.path.splitext(output_md)[0] + ".xlsx"
        generate_excel(statements, output_xlsx)

    print_summary(statements)


def cmd_regenerate():
    """从已有的 JSON 重新生成 Markdown 报告（可选附加 Excel）"""
    if len(sys.argv) < 3:
        print("用法: python3 doc_fact_check.py --regenerate <checklist_result.json> [输出报告.md] [--excel]")
        sys.exit(1)

    flags      = [a for a in sys.argv[1:] if a.startswith("--")]
    want_excel = "--excel" in flags
    pos_args   = [a for a in sys.argv[2:] if not a.startswith("--")]

    json_path = pos_args[0]
    output_md = pos_args[1] if len(pos_args) > 1 else ""

    with open(json_path, 'r', encoding='utf-8') as f:
        statements = json.load(f)
    print(f"从 {json_path} 加载 {len(statements)} 条表述")

    base_dir = os.path.dirname(os.path.abspath(json_path))
    doc_name = os.path.splitext(os.path.basename(json_path))[0].replace("checklist_result", "核对清单")
    if not output_md:
        output_md = os.path.join(base_dir, f"{doc_name}.md")
    elif not os.path.isabs(output_md):
        output_md = os.path.join(base_dir, output_md)

    generate_markdown(statements, output_md, doc_name=doc_name)

    if want_excel:
        output_xlsx = os.path.splitext(output_md)[0] + ".xlsx"
        generate_excel(statements, output_xlsx)

    print_summary(statements, stage="人工复查后重生成")


def main():
    if len(sys.argv) >= 2 and sys.argv[1] == "--regenerate":
        cmd_regenerate()
    elif len(sys.argv) >= 2 and sys.argv[1] in ("--help", "-h"):
        print(__doc__)
    else:
        cmd_full_check()


if __name__ == "__main__":
    main()
